import time
import logging
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Start time
st = time.time()

# Company reviews to be scraped and its total pages
company = input("Hi, please enter the company name to get reviews: ")
if company == "":
    company = "www"

tot_pages = int(input("Enter the total number of pages the website has or enter -1: "))
if tot_pages == -1:
    tot_pages = 1

# enter explicit delay to prevent site from banning
delay = int(input("Enter Delay in seconds whilescraping between two pages: "))

# Configure logging
logging.basicConfig(filename='scraping.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# def rating_converter(str):
#     if str == "-":
#         return 0
#     return float(review.rating)

class Review:
    def __init__(self, name, location, title, rating, date, description):
        self.name = name
        self.location = location
        self.title = title
        self.rating = rating
        self.date = date
        self.description = description

    def __str__(self):
        return (f"Name: {self.name}, Location: {self.location}, Title: {self.title}, "
                f"Rating: {self.rating}, Date: {self.date}, Description: {self.description}")

def extract_review_elements(soup):
    def get_names():
        name_blocks = soup.find_all("div", class_="row-inline align-center mb16px-desktop")
        names = []
        for block in name_blocks:
            name_span_tag = block.find('span')
            if name_span_tag and name_span_tag.text:
                names.append(name_span_tag.text)
            else:
                name_a_tag = block.find('a')
                names.append(name_a_tag.text if name_a_tag else "-")
        return names

    def get_locations():
        location_blocks = soup.find_all("div", class_="row-inline align-center mb16px-desktop")
        locations = []
        for location_div in location_blocks:
            location = location_div.find('span', class_='location-line grey-text row-inline align-center flex-wrap-wrap text-overflow-ellipsis')
            if location:
                locations.append(location.text.strip())
            else:
                locations.append("-")
        return locations

    def get_titles():
        title_blocks = soup.find_all("div", class_="f-component-info-header")
        titles = []
        for block in title_blocks:
            title_tag = block.find('span')
            if title_tag and title_tag.text:
                titles.append(title_tag.text)
            else:
                h2_tag = block.find('h2')
                titles.append(h2_tag.text if h2_tag else "-")
        return titles

    def get_ratings():
        rating_blocks = soup.find_all("div", class_="row-inline mb24px-desktop")
        ratings = []
        for rating_div in rating_blocks:
            rating = rating_div.find('div', class_='rating-title action-element bold-link-third')
            if rating:
                ratings.append(rating.text.strip())
            else:
                ratings.append("-")
        return ratings

    def get_dates():
        date_blocks = soup.find_all("time", class_="mr24px-desktop")
        return [date_block.get("datetime")[:10].replace("-", "/") for date_block in date_blocks]

    def get_descriptions():
        description_blocks = soup.find_all("div", class_="overflow-text")
        return [block.text for block in description_blocks]

    return (get_names(), get_locations(), get_titles(), get_ratings(), get_dates(), get_descriptions())

def get_reviews_from_page(company, page):
    url = f'https://{company}.pissedconsumer.com/review.html?page={page}'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'lxml')

    if response.status_code != 200:
        print("Wrong input provided. Exiting...")
        exit()

    names, locations, titles, ratings, dates, descriptions = extract_review_elements(soup)

    logging.info(f"Page {page} scraped...")
    print(f"Page {page} scraped...")

    reviews = []
    for i in range(max(len(names), len(locations), len(titles), len(ratings), len(dates), len(descriptions))):
        review = Review(
            names[i],
            locations[i] if i < len(locations) else "-",
            titles[i] if i < len(titles) else "-",
            ratings[i] if i < len(ratings) else "-",
            dates[i] if i < len(dates) else "-",
            descriptions[i] if i < len(descriptions) else "-"
        )
        reviews.append(review)

    return reviews

def fetch_all_reviews(company, max_pages):
    print("Scraping started...")
    all_reviews = []
    for page in range(1, max_pages + 1):
        reviews = get_reviews_from_page(company, page)
        all_reviews.extend(reviews)
        # Add a sleep to avoid being rate limited
        time.sleep(delay)

    logging.info(f"All reviews fetched. Total pages: {max_pages}")
    print(f"\nAll reviews fetched. Total pages: {max_pages}")
    return all_reviews

# Fetch reviews from the specified pages
reviews = fetch_all_reviews(company, tot_pages)

# Convert reviews data to a DataFrame
columns = ['Sr. No', 'Name', 'Location', 'Review Description', 'Rating', 'Review Comment Date', 'Issue With', 'Parameters']
reviews_data = [(review.name, review.location, review.description, review.rating, review.date, review.title, "", "") for review in reviews]
df = pd.DataFrame(reviews_data, columns=columns)

# Create Excel Workbook and Writer
excel_file = f'{company}_reviews.xlsx'
workbook = Workbook()
sheet = workbook.active

# Write headers
for col_num, header in enumerate(columns, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write data rows
for row_num, review in enumerate(reviews, 2):
    sheet.cell(row=row_num, column=1, value=row_num - 1)  # Sr. No
    sheet.cell(row=row_num, column=2, value=review.name)
    sheet.cell(row=row_num, column=3, value=review.location)
    sheet.cell(row=row_num, column=4, value=review.description)
    sheet.cell(row=row_num, column=5, value=review.rating) # user should use rating_converter(review.rating) if they wish to have a float value, non rated will have 0 which can break aggregate fxns in excel outputing data ()
    sheet.cell(row=row_num, column=6, value=review.date)
    sheet.cell(row=row_num, column=7, value=review.title)
    sheet.cell(row=row_num, column=8, value="-")

# Set column widths
column_widths = [7, 20, 25, 50, 7, 22, 30, 15]
for i, width in enumerate(column_widths):
    sheet.column_dimensions[chr(65 + i)].width = width

# Make headers bold
for cell in sheet[1]:
    cell.font = Font(bold=True)

# Save the file
workbook.save(filename=excel_file)

# End time
et = time.time()

print(f'Excel file "{excel_file}" created successfully.')
logging.info(f"\nTotal reviews collected: {len(reviews)}")
logging.info(f"\nTotal execution time for the program: {et - st:.2f} seconds")
print(f"\nTotal reviews collected: {len(reviews)}")
print(f"Total execution time for the program: {et - st:.2f} seconds")
print("\nScraping completed.")
logging.info("Scraping completed.")
